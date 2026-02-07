from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from decimal import Decimal
from django.db.models import Sum


class StampExcelService:
    """
    Service for generating Excel exports of stamp calculations.
    Provides formatted spreadsheets with proper styling and formatting.
    """

    @staticmethod
    def export_basic_report(queryset):
        """
        Export basic Excel report with stamp data.
        Includes company, date, value, copies, rate, and total.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stamp Report"

        # Headers
        headers = [
            "الشركة",
            "تاريخ المطالبه",
            "قيمة الأعمال",
            "عدد النسخ",
            "النسبة",
            "إجمالي الدمغة",
        ]
        ws.append(headers)

        # Add data rows
        for s in queryset:
            ws.append(
                [
                    s.company.name,
                    s.invoice_date.strftime("%Y-%m-%d") if s.invoice_date else "—",
                    s.value_of_work,
                    s.invoice_copies,
                    s.stamp_rate,
                    s.d1,
                ]
            )

        # Adjust column width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_formatted_report(queryset):
        """
        Export Excel report with professional formatting.
        Includes styling, borders, colors, and summary row.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "تقرير الدمغة"

        # Define styles
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        data_font = Font(name="Arial", size=11)
        data_alignment = Alignment(horizontal="right", vertical="center")

        border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        total_font = Font(name="Arial", size=12, bold=True)
        total_fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )

        # Headers
        headers = [
            "الشركة",
            "تاريخ المطالبه",
            "قيمة الأعمال",
            "عدد النسخ",
            "النسبة",
            "إجمالي الدمغة",
        ]

        # Write and style headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style

        # Write data rows
        row_num = 2
        total_amount = Decimal("0")

        for s in queryset:
            ws.cell(row=row_num, column=1, value=s.company.name)
            ws.cell(
                row=row_num,
                column=2,
                value=s.invoice_date.strftime("%Y-%m-%d") if s.invoice_date else "—",
            )
            ws.cell(row=row_num, column=3, value=float(s.value_of_work))
            ws.cell(row=row_num, column=4, value=s.invoice_copies)
            ws.cell(row=row_num, column=5, value=float(s.stamp_rate))
            ws.cell(row=row_num, column=6, value=float(s.d1))

            total_amount += s.d1

            # Apply styling to data rows
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border_style

            row_num += 1

        # Add total row
        ws.cell(row=row_num, column=5, value="الإجمالي:")
        ws.cell(row=row_num, column=6, value=float(total_amount))

        # Style total row
        for col_num in [5, 6]:
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

        # Adjust column widths
        column_widths = {
            "A": 30,  # الشركة
            "B": 18,  # تاريخ المطالبه
            "C": 18,  # قيمة الأعمال
            "D": 15,  # عدد النسخ
            "E": 15,  # النسبة
            "F": 20,  # إجمالي الدمغة
        }

        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    @staticmethod
    def export_company_summary_report(queryset):
        """
        Export Excel report grouped by company with subtotals.
        Shows summary statistics per company.
        """

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ملخص الشركات"

        # Styles
        header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        data_font = Font(name="Arial", size=11)
        data_alignment = Alignment(horizontal="right", vertical="center")

        border_style = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Headers
        headers = ["الشركة", "عدد الفواتير", "إجمالي النسخ", "إجمالي الدمغة"]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style

        # Get grouped data
        grouped_data = (
            queryset.values("company__name", "company_id")
            .annotate(
                total_amount=Sum("d1"),
                total_copies=Sum("invoice_copies"),
                invoice_count=Sum("id"),
            )
            .order_by("-total_amount")
        )

        # Write data
        row_num = 2
        grand_total = Decimal("0")
        total_copies = 0
        total_invoices = 0

        for item in grouped_data:
            ws.cell(row=row_num, column=1, value=item["company__name"])
            ws.cell(
                row=row_num,
                column=2,
                value=queryset.filter(company_id=item["company_id"]).count(),
            )
            ws.cell(row=row_num, column=3, value=item["total_copies"] or 0)
            ws.cell(row=row_num, column=4, value=float(item["total_amount"] or 0))

            grand_total += Decimal(str(item["total_amount"] or 0))
            total_copies += item["total_copies"] or 0
            total_invoices += queryset.filter(company_id=item["company_id"]).count()

            # Apply styling
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border_style

            row_num += 1

        # Add grand total row
        total_font = Font(name="Arial", size=12, bold=True)
        total_fill = PatternFill(
            start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"
        )

        ws.cell(row=row_num, column=1, value="الإجمالي الكلي")
        ws.cell(row=row_num, column=2, value=total_invoices)
        ws.cell(row=row_num, column=3, value=total_copies)
        ws.cell(row=row_num, column=4, value=float(grand_total))

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border_style

        # Adjust column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 22

        # Freeze header row
        ws.freeze_panes = "A2"

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
