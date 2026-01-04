from decimal import Decimal
from django.utils.dateparse import parse_date
from stamps.admin import format_millions
from stamps.models import StampCalculation, Company
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.enums import TA_RIGHT
from io import BytesIO
import os
import openpyxl
from openpyxl.utils import get_column_letter
from arabic_reshaper import reshape
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from bidi.algorithm import get_display
from django.conf import settings
import io
from datetime import date,timedelta,datetime
import textwrap
from site_settings.models import SiteConfiguration
from decimal import Decimal
from django.db.models import Sum, Q
from django.utils.dateparse import parse_date
from django.utils import timezone
from typing import Optional
from django.db.models.functions import TruncYear


class StampService:
    """
    Business logic for StampCalculation with improved error handling,
    performance optimizations, and better separation of concerns.
    """

    PREVIOUS_YEAR_MULTIPLIER = Decimal("0.7")
    PENSION_MULTIPLIER = Decimal("0.2")
    MONTHS_PER_YEAR = 12

    def __init__(self, retired_engineers: Optional[int] = None):
        if retired_engineers is None:
            config = SiteConfiguration.objects.only(
                "number_of_retired_engineers"
            ).first()
            retired_engineers = (
                getattr(config, "number_of_retired_engineers", 0) if config else 0
            )

        self.retired_engineers = retired_engineers or 0
        self.current_year = timezone.now().year

    @staticmethod
    def get_queryset():
        """Get optimized base queryset with related data."""
        return StampCalculation.objects.select_related("company")

    @staticmethod
    def get_last_year(date_to):
        date_str = date_to[0]
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d").date()
            last_year = date_obj.year
        except:
            return None

        return last_year

    @classmethod
    def get_filtered_queryset(
        cls,
        company_id: Optional[int] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
        sort: str = "-created_at",
    ):
        queryset = cls.get_queryset()
        queryset = cls.filter(queryset, company_id, date_from, date_to)
        return cls.sort(queryset, sort)

    @staticmethod
    def filter(queryset, company_id=None, date_from=None, date_to=None,user=None):
        filters = Q()

        if company_id and str(company_id).lower() not in ["none", ""]:
            filters &= Q(company_id=company_id)

        if date_from:
            parsed_date = parse_date(date_from)
            if parsed_date:
                filters &= Q(invoice_date__gte=parsed_date)

        if date_to:
            parsed_date = parse_date(date_to)
            if parsed_date:
                filters &= Q(invoice_date__lte=parsed_date)

        if user:
            filters &= Q(user=user)

        return queryset.filter(filters) if filters else queryset

    @staticmethod
    def filter_by_years(queryset, years: int | None):
        if not years:
            return queryset

        start_date = timezone.now().date() - timedelta(days=365 * years)
        return queryset.filter(invoice_date__gte=start_date)

    @staticmethod
    def sort(queryset, sort: str = "-created_at"):
        allowed_sorts = ["invoice_date", "-invoice_date", "created_at", "-created_at"]
        return queryset.order_by(sort if sort in allowed_sorts else "-created_at")

    @staticmethod
    def total_amount(queryset) -> Decimal:
        """Calculate total amount from d1 field."""
        result = queryset.aggregate(total=Sum("d1"))["total"]
        return Decimal(str(result)) if result else Decimal("0")

    def _total_for_previous_year(
        self, queryset, current_year: Optional[int] = None
    ) -> Decimal:

        year = current_year if current_year is not None else self.current_year
        previous_year = year - 1

        stamps = queryset.filter(invoice_date__year=previous_year)
        total = stamps.aggregate(total=Sum("d1"))["total"]

        if not total:
            return Decimal("0")

        return Decimal(str(total)) * self.PREVIOUS_YEAR_MULTIPLIER

    def calculate_pension(
        self, queryset, year ,current_year: Optional[int] = None
    ) -> Decimal:

        if year:
            year = year 
        else:
            year = current_year if current_year is not None else self.current_year

        current_total = self.total_amount(queryset)
        previous_total = self._total_for_previous_year(queryset, year)

        pension = ((current_total * self.PENSION_MULTIPLIER) + previous_total) / (
            self.retired_engineers * self.MONTHS_PER_YEAR
        )

        return pension

    @staticmethod
    def total_companies(queryset) -> int:
        return queryset.values("company_id").distinct().count()

    @staticmethod
    def total_amount_for_company(queryset, company_id: int) -> Decimal:
        result = queryset.filter(company_id=company_id).aggregate(total=Sum("d1"))[
            "total"
        ]
        return Decimal(str(result)) if result else Decimal("0")

    @staticmethod
    def grouped_by_company(queryset):
        return (
            queryset.values("company__name", "company_id", "stamp_rate")
            .annotate(total=Sum("d1"))
            .order_by("-total")
        )

    @staticmethod
    def yearly_chart(queryset):
        stamps = (
            queryset.filter(invoice_date__isnull=False)
            .annotate(year=TruncYear("invoice_date"))
            .values("year")
            .annotate(total=Sum("d1"))
            .order_by("year")
        )

        categories = []
        yearly = []
        cumulative = []

        running_total = 0

        for item in stamps:
            year = item["year"].strftime("%Y")
            value = round(float(item["total"]), 2)

            categories.append(year)
            yearly.append(value)

            running_total += value
            cumulative.append(round(running_total, 2))

        yearly = [format_millions(v) for v in yearly]
        cumulative = [format_millions(v) for v in cumulative]

        return {
            "categories": categories,
            "yearly": yearly,
            "cumulative": cumulative,
        }

    @staticmethod
    def create_from_form(form,user):
        new_company_name = form.cleaned_data.get("new_company_name", "").strip()
        company = form.cleaned_data.get("company")

        if new_company_name:
            company, created = Company.objects.get_or_create(
                name__iexact=new_company_name, defaults={"name": new_company_name}
            )

        if not company:
            raise ValueError("Either company or new_company_name must be provided")

        stamp = form.save(commit=False)
        stamp.company = company
        stamp.user = user
        stamp.save()

        return stamp

    @staticmethod
    def fix_arabic(text):
        if text is None:
            return ""
        return get_display(reshape(str(text)))

    @staticmethod
    def export_pdf(queryset):

        buffer = BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20,
        )

        FONT_PATH = os.path.join("static", "fonts", "Amiri-Bold.ttf")
        try:
            pdfmetrics.registerFont(TTFont("Amiri", FONT_PATH))
            font_name = "Amiri"
        except:
            font_name = "Helvetica"

        styles = getSampleStyleSheet()

        arabic_style = ParagraphStyle(
            name="Arabic",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=12,
            alignment=TA_RIGHT,
            leading=14,
        )

        number_style = ParagraphStyle(
            name="Number",
            parent=arabic_style,
            alignment=TA_RIGHT,
        )

        # ----------- NEW: Calculate TOTAL -----------
        total_amount = StampService.total_amount(queryset)

        total_paragraph = Paragraph(
            StampService.fix_arabic(
                f"إجمالي الدمغة لكل الشركات بالمليون: {total_amount:,} جنيه مصري"
            ),
            arabic_style,
        )
        # --------------------------------------------

        # RTL: REVERSED column order
        headers = [
            StampService.fix_arabic("إجمالي الدمغة"),
            StampService.fix_arabic("النسبة"),
            StampService.fix_arabic("عدد النسخ"),
            StampService.fix_arabic("قيمة الأعمال"),
            StampService.fix_arabic("تاريخ المطالبه"),
            StampService.fix_arabic("الشركة"),
        ]

        table_data = [[Paragraph(h, arabic_style) for h in headers]]

        for s in queryset:
            row = [
                Paragraph(f"{s.d1:,}", number_style),
                Paragraph(str(s.stamp_rate), number_style),
                Paragraph(str(s.invoice_copies), number_style),
                Paragraph(f"{s.value_of_work:,}", number_style),
                Paragraph(
                    s.invoice_date.strftime("%Y-%m-%d") if s.invoice_date else "—",
                    number_style,
                ),
                Paragraph(StampService.fix_arabic(s.company.name), arabic_style),
            ]

            table_data.append(row)

        table = Table(
            table_data,
            colWidths=[95, 60, 60, 90, 80, 120],
            repeatRows=1,
        )

        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgreen),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))

        # Build PDF with total on top
        doc.build([total_paragraph, Spacer(1, 12), table])

        pdf = buffer.getvalue()
        buffer.close()
        return pdf

    @staticmethod
    def export_to_pdf_for_spacific_company(queryset, company_id):

        company = Company.objects.get(id=company_id)
        buffer = io.BytesIO()

        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        # ================= Fonts ================= #
        pdfmetrics.registerFont(
            TTFont(
                "Amiri",
                settings.BASE_DIR / "static/fonts/Amiri-Regular.ttf"
            )
        )
        pdfmetrics.registerFont(
            TTFont(
                "Amiri-Bold",
                settings.BASE_DIR / "static/fonts/Amiri-Bold.ttf"
            )
        )

        # ================= Layout constants ================= #
        HEADER_FONT = ("Amiri", 12)
        TITLE_FONT = ("Amiri-Bold", 14)
        PARA_FONT = ("Amiri", 11)
        TABLE_HEADER_FONT = ("Amiri-Bold", 11)
        TABLE_ROW_FONT = ("Amiri", 11)

        LINE_HEIGHT = 0.8 * cm
        ROW_HEIGHT = 0.7 * cm
        LEFT = 2 * cm
        RIGHT = width - 2 * cm
        TOP_MARGIN = 5 * cm
        FOOTER_LEFT = 5 * cm
        y = height - TOP_MARGIN

        # ================= Header ================= #
        c.setFont(*HEADER_FONT)
        c.drawRightString(
            RIGHT,
            y,
            StampService.fix_arabic(f" القاهرة في : {date.today().strftime('%Y-%m-%d')}")
        )

        y -= 1.2 * cm
        c.drawRightString(RIGHT, y, StampService.fix_arabic(f"السادة شركة / {company.name}"))

        # "تحية طيبة وبعد"
        y -= 1.5 * cm
        c.setFont("Amiri-Bold", 13)
        c.drawCentredString(width / 2, y, StampService.fix_arabic("تحية طيبة و بعد"))

        # Title
        y -= 1.2 * cm
        c.setFont(*TITLE_FONT)
        c.drawCentredString(width / 2, y, StampService.fix_arabic("مطالبة نموذج رقم ( 1 )"))

        # ================= Intro paragraph ================= #
        y -= 2 * cm
        c.setFont(*PARA_FONT)

        paragraph_text = (
            "بمراجعة حجم اعمالكم وجد لديكم كدمغة هندسية لم تورد طبقا للقانون رقم 66 لسنة 1974 "
            "و مواد الدمغة ارقام 45&46&47&48&99 و مواد اللائحة التنفيذية ارقام 130&131&132&149 "
            "وحكم الدستورية في الدعوي رقم 16 لسنة 38 ق دستورية في 5-12-2020. "
            "وحكم جنح مستانف سمالوط رقم 206لسنة 2018"
        )

        # WRAP paragraph across full width (RIGHT → LEFT)
        max_width = RIGHT - LEFT
        wrapped_lines = textwrap.wrap(paragraph_text, width=85)

        for line in wrapped_lines:
            c.drawRightString(RIGHT, y, StampService.fix_arabic(line))
            y -= 0.7 * cm

        # ================= Table ================= #
        y -= 1.2 * cm

        headers = [
            "التاريخ",
            "حجم الأعمال",
            "الدمغة الهندسية",
            "عدد النسخ",
            "إجمالي الدمغة",
        ]

        col_widths = [
            3 * cm,
            4 * cm,
            3.5 * cm,
            3 * cm,
            4.5 * cm,
        ]

        c.setFont(*TABLE_HEADER_FONT)
        x = RIGHT

        c.line(LEFT, y + 0.4 * cm, RIGHT, y + 0.4 * cm)

        for header, w in zip(headers, col_widths):
            c.drawRightString(x, y, StampService.fix_arabic(header))
            x -= w

        c.line(LEFT, y - 0.3 * cm, RIGHT, y - 0.3 * cm)

        y -= LINE_HEIGHT

        # ---------- Table rows ---------- #
        total = 0
        c.setFont(*TABLE_ROW_FONT)

        for obj in queryset:
            x = RIGHT

            row = [
                obj.invoice_date.strftime("%Y-%m-%d") if obj.invoice_date else "—",
                f"{obj.value_of_work:,}",
                f"{obj.stamp_rate}",
                obj.invoice_copies,
                f"{obj.d1:,}",
            ]

            total += obj.d1

            for value, w in zip(row, col_widths):
                c.drawRightString(x, y, StampService.fix_arabic(str(value)))
                x -= w

            c.setStrokeColorRGB(0.85, 0.85, 0.85)
            c.line(LEFT, y - 0.2 * cm, RIGHT, y - 0.2 * cm)
            c.setStrokeColorRGB(0, 0, 0)

            y -= ROW_HEIGHT

            if y < 4 * cm:
                c.showPage()
                c.setFont(*TABLE_ROW_FONT)
                y = height - 2 * cm

        # ================= Total ================= #
        y -= 0.5 * cm
        c.setFont("Amiri-Bold", 12)

        c.line(LEFT, y + 0.4 * cm, RIGHT, y + 0.4 * cm)

        c.drawRightString(
            RIGHT,
            y,
            StampService.fix_arabic(f"الإجمالي : {total:,} جنيه مصري")
        )

        # ================= Footer ================= #
        last_points = [
            "١- يمكنكم الطعن و التقدم بمستنداتكم خلال خمسه عشر  يوما من تاريخ الاستلام لإدارة الدمغة بالنقابة العامة للمهندسين.. مع اثبات تسليم الطعن.",
            "٢- عدم الاعتراض أو الطعن خلال ١٥ يوم يعتبر مصادقة علي المديونية.",
            "٣- تطبق المادة ٩٩ عند التقاعس مدة ثلاث شهور عن الدفع.",
        ]

        y -= 1.6 * cm
        c.setFont("Amiri", 10)

        for point in last_points:
            c.drawRightString(RIGHT, y, StampService.fix_arabic(point))
            y -= 0.7 * cm   # المسافة بين النقاط

        y -= 1.2 * cm
        c.setFont("Amiri", 11)
        c.drawCentredString(FOOTER_LEFT, y, StampService.fix_arabic("وتفضلوا بقبول فائق الاحترام"))

        y -= 1.3 * cm
        c.setFont("Amiri-Bold", 11)
        c.drawCentredString(FOOTER_LEFT, y, StampService.fix_arabic("أمين الصندوق"))

        y -= 0.9 * cm
        c.drawCentredString(FOOTER_LEFT, y, StampService.fix_arabic("د / معتز طلبة"))

        c.showPage()
        c.save()

        buffer.seek(0)
        return buffer.getvalue()

    @staticmethod
    def export_excel(queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stamp Report"

        headers = ["الشركة", "تاريخ المطالبه", "قيمة الأعمال", "عدد النسخ", "النسبة", "إجمالي الدمغة"]
        ws.append(headers)

        for s in queryset:
            ws.append(
                [
                    s.company.name,
                    s.invoice_date.strftime("%Y-%m-%d") if s.invoice_date else "—",
                    s.value_of_work,
                    s.invoice_copies,
                    s.stamp_rate,
                    s.d1,
                ]
            )

        # Adjust column width
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

        output = BytesIO()
        wb.save(output)
        return output.getvalue()
